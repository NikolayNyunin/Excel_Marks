import sys
import os
import time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyQt5.QtWidgets import QApplication, QWidget, QFileDialog, QPushButton, QLineEdit, QTextEdit, QLabel, QGridLayout
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont


def is_number(mark):
    if (''.join(mark.split('.'))).isdigit():
        return True
    return False


def get_needed_mark(mark):
    if not is_number(mark):
        return mark
    else:
        mark = float(mark)

    if mark == 0:
        return '0'
    elif mark < 2.5:
        return '2'
    elif mark < 3.5:
        return '3'
    elif mark < 4.5:
        return '4'
    return '5'


class ExcelMarksInterface(QWidget):
    def __init__(self):
        super().__init__()
        self.needed_file_description, self.select_file_button, self.selected_file_label, self.period_label,\
            self.period_input, self.form_data_description, self.form_input, self.start_analysing_button,\
            self.or_label, self.start_analysing_all_button, self.output_console = [None] * 11
        self.init_ui()
        self.analyser = ExcelMarksAnalyser()
        self.filename = None
        self.show()

    def init_ui(self):
        self.setFixedSize(800, 600)
        self.setWindowTitle('Обработка оценок учеников')

        grid = QGridLayout()
        grid.setContentsMargins(40, 20, 40, 30)
        grid.setSpacing(15)
        self.setLayout(grid)

        self.needed_file_description = QLabel('Выберите файл с итоговыми оценками:', self)
        self.needed_file_description.setFont(QFont('Arial', 13))
        grid.addWidget(self.needed_file_description, 0, 0, 1, 3, alignment=Qt.AlignCenter)

        self.select_file_button = QPushButton('Выбрать файл', self)
        self.select_file_button.setFont(QFont('Arial', 12))
        self.select_file_button.clicked.connect(self.select_file)
        self.select_file_button.setFixedSize(150, 40)
        grid.addWidget(self.select_file_button, 1, 0, alignment=Qt.AlignCenter)

        self.selected_file_label = QLabel('Файл не выбран.', self)
        self.selected_file_label.setFont(QFont('Arial', 12))
        grid.addWidget(self.selected_file_label, 1, 1, 1, 2)

        self.period_label = QLabel('Номер триместра или полугодия:', self)
        self.period_label.setFont(QFont('Arial', 12))
        grid.addWidget(self.period_label, 2, 0, 1, 2, alignment=Qt.AlignCenter)

        self.period_input = QLineEdit('', self)
        self.period_input.setFont(QFont('Arial', 14))
        grid.addWidget(self.period_input, 2, 2)

        self.form_data_description = QLabel('Введите полное название класса (разделяя номер и букву дефисом):', self)
        self.form_data_description.setFont(QFont('Arial', 13))
        self.form_data_description.setFixedWidth(570)
        grid.addWidget(self.form_data_description, 3, 0, 1, 2, alignment=Qt.AlignCenter)

        self.form_input = QLineEdit('', self)
        self.form_input.setFont(QFont('Arial', 14))
        self.form_input.setMaximumWidth(150)
        grid.addWidget(self.form_input, 3, 2, alignment=Qt.AlignRight)

        self.start_analysing_button = QPushButton('Обработать', self)
        self.start_analysing_button.setFont(QFont('Arial', 13))
        self.start_analysing_button.clicked.connect(self.analyse)
        self.start_analysing_button.setAutoDefault(True)
        self.start_analysing_button.setFixedSize(200, 50)
        grid.addWidget(self.start_analysing_button, 4, 0, 1, 3, alignment=Qt.AlignCenter)

        self.or_label = QLabel('ИЛИ', self)
        self.or_label.setFont(QFont('Arial', 13))
        grid.addWidget(self.or_label, 5, 0, 1, 3, alignment=Qt.AlignCenter)

        self.start_analysing_all_button = QPushButton('Обработать все файлы в папке')
        self.start_analysing_all_button.setFont(QFont('Arial', 13))
        self.start_analysing_all_button.clicked.connect(self.analyse_all)
        self.start_analysing_all_button.setFixedSize(300, 50)
        grid.addWidget(self.start_analysing_all_button, 6, 0, 1, 3, alignment=Qt.AlignCenter)

        self.output_console = QTextEdit('', self)
        self.output_console.setFont(QFont('Arial', 12))
        self.output_console.setReadOnly(True)
        self.output_console.setMaximumHeight(300)
        grid.addWidget(self.output_console, 7, 0, 1, 3)

    def select_file(self):  # метод для отображения окна выбора файла
        filename = QFileDialog.getOpenFileName(self, 'Выбор файла для обработки')[0]
        if filename == '':
            self.selected_file_label.setText('Файл не выбран.')
        else:
            self.filename = filename
            self.selected_file_label.setText('Выбранный файл: "{}".'.format(filename.split('/')[-1]))

    def analyse(self, form=None, period=None):  # метод для обработки файла
        if self.filename in (None, ''):  # проверка на то, что файл не выбран
            self.output_console.append('Ошибка: Файл не выбран.\n')
            return
        if not self.filename.endswith('.xlsx'):  # проверка на расширение файла
            self.output_console.append('Ошибка: Неподдерживаемое расширение файла: "{}".\n'.
                                       format(self.filename.split('.')[-1]))
            return

        if not form:
            form = self.form_input.text().upper()
        if form == '':  # проверка на то, что класс не указан
            self.output_console.append('Ошибка: Класс не указан.\n')
            return
        if '-' not in form:  # проверка на отсутствие дефиса в названии класса
            self.output_console.append('Ошибка: Неправильный формат названия класса.\n')
            return

        if not period:
            period = self.period_input.text()
        if period == '':
            self.output_console.append('Ошибка: Период аттестации не указан.\n')
            return
        if not period.isdigit():
            self.output_console.append('Ошибка: Неправильный формат периода аттестации.\n')
            return
        if int(period) > 3 or int(period) < 1:
            self.output_console.append('Ошибка: Неправильный формат периода аттестации.\n')
            return

        try:
            start_time = time.time()

            new_filename = 'Заключение по итоговым оценкам_{}.xlsx'.format(form)

            self.analyser.analyse_file(self.filename, form, period)
            self.analyser.create_resulting_file(new_filename, form)  # , period)
            self.analyser.reset()

            self.output_console.append('Успешно обработано: "{}" ({}).'.format(self.filename, form))
            self.output_console.append('Файл "{}" успешно создан.'.format(new_filename))
            self.output_console.append('Длительность выполнения: {} сек.\n'.
                                       format(str(round(time.time() - start_time, 2))))

        except Exception as e:
            self.output_console.append('Ошибка: {}.\n'.format(e))

    def analyse_all(self):
        if self.filename in (None, ''):  # проверка на то, что файл не выбран
            self.output_console.append('Ошибка: Файл не выбран.\n')
            return
        if not self.filename.endswith('.xlsx'):  # проверка на расширение файла
            self.output_console.append('Ошибка: Неподдерживаемое расширение файла: "{}".\n'.
                                       format(self.filename.split('.')[-1]))
            return

        period = self.period_input.text()
        if period == '':
            self.output_console.append('Ошибка: Период аттестации не указан.\n')
            return
        if not period.isdigit():
            self.output_console.append('Ошибка: Неправильный формат периода аттестации.\n')
            return
        if int(period) > 3 or int(period) < 1:
            self.output_console.append('Ошибка: Неправильный формат периода аттестации.\n')
            return

        if len(self.filename.split('/')) > 1:
            path = '/'.join(self.filename.split('/')[:-1]) + '/'
        else:
            path = ''

        files = os.listdir(path)
        forms = []
        for file in files:
            file = file.split('-')
            if not file[-1].endswith('.xlsx'):
                continue

            for i in range(len(file) - 1):
                if len(file[i]) < 1 or len(file[i + 1]) < 1:
                    continue

                if file[i + 1][0].isalpha():
                    if file[i].endswith('10') or file[i].endswith('11'):
                        forms.append(file[i][-2:] + '-' + file[i + 1][0])
                        break
                    elif file[i][-1].isdigit():
                        forms.append(file[i][-1] + '-' + file[i + 1][0])
                        break

        for form in forms:
            self.analyse(form, period)
            self.output_console.repaint()

        self.output_console.append('Обработка файлов завершена.\n')


class ExcelMarksAnalyser:
    def __init__(self):
        self.all_subjects = None
        self.students = {}
        self.THIN = Side(border_style='thin', color='000000')
        self.THICK = Side(border_style='thick', color='000000')
        self.DOUBLE = Side(border_style='double', color='000000')

    def reset(self):
        self.all_subjects = None
        self.students = {}

    def get_average_marks(self, path, filenames):  # метод для получения средних баллов из первого файла
        if len(filenames) == 1:
            file_num = 0
        else:
            raise ValueError('Слишком много файлов со средними оценками')

        workbook = load_workbook('{}{}'.format(path, filenames[file_num]), read_only=True)
        sheet = workbook.active

        subjects = list(map(lambda s: s.value, sheet[6][1:]))
        self.all_subjects = subjects.copy()

        row_num = 7
        while True:  # пробегаемся по всем рядам таблицы с учениками
            if row_num > sheet.max_row:
                break
            row = list(map(lambda c: c.value, sheet[row_num]))
            student = row[0]
            if student in ('', None):  # проверка на пустоту ячейки
                break
            if student not in self.students.keys():
                self.students[student] = {}

            for mark_index in range(len(row[1:])):  # пробегаемся по всем оценкам данного ученика
                mark = row[1:][mark_index]
                subject = subjects[mark_index]

                if subject not in self.students[student].keys():
                    self.students[student][subject] = [[None] * 2, [None] * 2, [None] * 2]

                self.students[student][subject][file_num][0] = mark

            row_num += 1

    def get_final_marks(self, filename, form):  # метод для получения триместровых оценок
        workbook = load_workbook(filename, read_only=False)

        form_num = form.split('-')[0]
        sheet = workbook[form_num]

        data = list(map(lambda el: el.value, sheet['B']))
        index = data.index(form) + 1

        subj_index = index + 2
        col = 2
        subjects, periods = {}, {}
        while sheet.cell(row=subj_index + 1, column=col).value not in (None, ''):  # анализируем шапку
            subject = sheet.cell(row=subj_index, column=col).value
            if subject not in (None, ''):
                subjects[col] = subject

                if subject not in self.all_subjects:
                    self.all_subjects.append(subject)
            else:
                subjects[col] = subjects[col - 1]

            period = sheet.cell(row=subj_index + 1, column=col).value
            if period not in (None, ''):
                periods[col] = period

            col += 1

        max_col = col

        student_index = subj_index + 2
        student_name = sheet.cell(row=student_index, column=1).value
        while student_name not in (None, ''):  # пробегаемся по всем ученикам нужного класса
            short_name = ' '.join(student_name.split()[:2])

            if short_name not in self.students.keys():
                self.students[short_name] = {}

            for col in range(2, max_col):  # пробегаемся по всем оценкам данного ученика
                mark = sheet.cell(row=student_index, column=col).value
                if mark in (None, ''):
                    continue

                subject = subjects[col]
                if subject not in self.students[short_name].keys():
                    self.students[short_name][subject] = [[None] * 2, [None] * 2, [None] * 2]

                if int(form_num) in (10, 11):
                    if periods[col] == 'Первое полугодие':  # если это итоговая 1 полугодия
                        self.students[short_name][subject][0][1] = mark
                    elif periods[col] == 'Второе полугодие':  # если это итоговая 2 полугодия
                        self.students[short_name][subject][1][1] = mark
                else:
                    if periods[col] == '1 триместр':  # если это итоговая 1 триместра
                        self.students[short_name][subject][0][1] = mark
                    elif periods[col] == '2 триместр':  # если это итоговая 2 триместра
                        self.students[short_name][subject][1][1] = mark
                    elif periods[col] == '3 триместр':  # если это итоговая 3 триместра
                        self.students[short_name][subject][2][1] = mark

            student_index += 1
            student_name = sheet.cell(row=student_index, column=1).value

    def analyse_file(self, filename, form, period):  # основной метод для обработки данных файлов
        if len(filename.split('/')) > 1:
            path = '/'.join(filename.split('/')[:-1]) + '/'
        else:
            path = ''

        filenames = []
        for file in os.listdir(path):
            if form in file and '.xlsx' in file:
                if 'I' not in file:
                    filenames.append(file)
                    break
                filenames.append(file)
        filenames.sort(key=lambda el: el.count('I'))
        if len(filenames) == 0:
            raise ValueError('Файл со средними оценками не найден')

        self.get_average_marks(path, filenames)
        self.get_final_marks(filename, form)

    def create_resulting_file(self, filename, form):  # метод для создания результирующего файла
        wrong_marks = []

        workbook = Workbook()
        sheet = workbook.active

        sheet['A1'] = 'Номер'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A1:A3')
        sheet['A1'].border, sheet['A2'].border, sheet['A3'].border = \
            Border(right=self.THIN), Border(right=self.THIN), Border(right=self.THIN, bottom=self.THIN)
        sheet.column_dimensions['A'].width = 8

        sheet['B1'] = 'Имя ученика'
        sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('B1:B3')
        sheet['B1'].border, sheet['B2'].border, sheet['B3'].border = \
            Border(right=self.THICK), Border(right=self.THICK), Border(right=self.THICK, bottom=self.THIN)
        sheet.column_dimensions['B'].width = 45

        student_index = 0  # порядковый номер ученика
        for student in sorted(self.students.keys()):  # пробегаемся по всем ученикам
            sheet.cell(row=student_index + 4, column=1).alignment = Alignment(horizontal='left')

            if student_index == len(self.students) - 1:
                sheet.cell(row=student_index + 4, column=2, value=student).border = Border(bottom=self.THIN,
                                                                                           right=self.THICK)
                sheet.cell(row=student_index + 4, column=1, value=student_index + 1).border = Border(bottom=self.THIN,
                                                                                                     right=self.THIN)
            else:
                sheet.cell(row=student_index + 4, column=1, value=student_index + 1).border = Border(right=self.THIN)
                sheet.cell(row=student_index + 4, column=2, value=student).border = Border(right=self.THICK)

            subject_index = 0
            for subject in sorted(self.all_subjects):  # пробегаемся по всем предметам

                subject_column = 3 + subject_index * 9
                if student_index == 0:  # если это список предметов первого ученика, заполняем шапку
                    sheet.cell(row=1, column=subject_column).value = subject
                    sheet.cell(row=1, column=subject_column).alignment = Alignment(horizontal='center')
                    sheet.merge_cells(start_row=1, start_column=subject_column,
                                      end_row=1, end_column=subject_column + 8)
                    sheet.cell(row=1, column=subject_column + 8).border = Border(right=self.THICK)

                for trimester in range(0, 3):  # пробегаемся по триместрам

                    column = subject_column + trimester * 3

                    if student_index == 0:  # если это список предметов первого ученика, заполняем шапку
                        if int(form.split('-')[0]) in (10, 11):
                            if trimester == 2:
                                sheet.cell(row=2, column=column).value = '-'
                            else:
                                sheet.cell(row=2, column=column).value = '{}-е полугодие'.format(trimester + 1)
                        else:
                            sheet.cell(row=2, column=column).value = '{}-й триместр'.format(trimester + 1)
                        sheet.cell(row=2, column=column).alignment = Alignment(horizontal='center')
                        sheet.merge_cells(start_row=2, start_column=column,
                                          end_row=2, end_column=column + 2)
                        sheet.cell(row=2, column=column).border = Border(top=self.THIN, bottom=self.THIN)
                        sheet.cell(row=2, column=column + 1).border = Border(top=self.THIN, bottom=self.THIN)

                        sheet.cell(row=3, column=column).value = 'ср. б.'
                        sheet.cell(row=3, column=column + 1).value = 'рек.'
                        sheet.cell(row=3, column=column + 2).value = 'фактич.'

                        sheet.cell(row=3, column=column).alignment = Alignment(horizontal='center')
                        sheet.cell(row=3, column=column + 1).alignment = Alignment(horizontal='center')
                        sheet.cell(row=3, column=column + 2).alignment = Alignment(horizontal='center')

                        sheet.cell(row=3, column=column).border = Border(bottom=self.THIN, right=self.THIN)
                        sheet.cell(row=3, column=column + 1).border = Border(bottom=self.THIN, right=self.THIN)

                        if trimester == 2:
                            sheet.cell(row=2, column=column + 2).border = Border(bottom=self.THIN,
                                                                                 top=self.THIN, right=self.THICK)
                            sheet.cell(row=3, column=column + 2).border = Border(bottom=self.THIN, right=self.THICK)
                        else:
                            sheet.cell(row=2, column=column + 2).border = Border(bottom=self.THIN,
                                                                                 top=self.THIN, right=self.DOUBLE)
                            sheet.cell(row=3, column=column + 2).border = Border(bottom=self.THIN, right=self.DOUBLE)

                    if subject in self.students[student].keys():  # если предмета нет в списке предметов ученика
                        marks = ['0' if mark is None else mark for mark in self.students[student][subject][trimester]]
                    else:
                        marks = ['0', '0']

                    if is_number(marks[0]):
                        marks[0] = float(marks[0])
                    if is_number(marks[1]):
                        marks[1] = int(marks[1])
                    sheet.cell(row=student_index + 4, column=column).value = marks[0]
                    sheet.cell(row=student_index + 4, column=column + 2).value = marks[1]

                    sheet.cell(row=student_index + 4, column=column).alignment = Alignment(horizontal='center')
                    sheet.cell(row=student_index + 4, column=column + 2).alignment = Alignment(horizontal='center')

                    recommended = get_needed_mark(str(marks[0]))
                    if is_number(recommended):
                        recommended = int(recommended)
                    sheet.cell(row=student_index + 4, column=column + 1).value = recommended
                    sheet.cell(row=student_index + 4, column=column + 1).alignment = Alignment(horizontal='center')

                    if recommended != marks[1]:
                        wrong_marks.append({'name': student, 'subject': subject, 'period': trimester,
                                            'average': marks[0], 'recommended': recommended, 'actual': marks[1]})

                        sheet.cell(row=student_index + 4, column=column).font = Font(b=True)
                        sheet.cell(row=student_index + 4, column=column + 1).font = Font(b=True)
                        sheet.cell(row=student_index + 4, column=column + 2).font = Font(b=True)
                        sheet.cell(row=student_index + 4, column=column).fill = PatternFill(
                            start_color='FF4040',
                            end_color='FF4040',
                            fill_type='solid')
                        sheet.cell(row=student_index + 4, column=column + 1).fill = PatternFill(
                            start_color='FF4040',
                            end_color='FF4040',
                            fill_type='solid')
                        sheet.cell(row=student_index + 4, column=column + 2).fill = PatternFill(
                            start_color='FF4040',
                            end_color='FF4040',
                            fill_type='solid')

                    if student_index == len(self.students) - 1:  # если это последний ученик в списке,
                        # рисуем нижнюю границу
                        sheet.cell(row=student_index + 4, column=column).border = Border(bottom=self.THIN,
                                                                                         right=self.THIN)
                        sheet.cell(row=student_index + 4, column=column + 1).border = Border(bottom=self.THIN,
                                                                                             right=self.THIN)

                        if trimester == 2:
                            sheet.cell(row=student_index + 4, column=column + 2).border = Border(bottom=self.THIN,
                                                                                                 right=self.THICK)
                        else:
                            sheet.cell(row=student_index + 4, column=column + 2).border = Border(bottom=self.THIN,
                                                                                                 right=self.DOUBLE)
                    else:
                        sheet.cell(row=student_index + 4, column=column).border = Border(right=self.THIN)
                        sheet.cell(row=student_index + 4, column=column + 1).border = Border(right=self.THIN)

                        if trimester == 2:
                            sheet.cell(row=student_index + 4, column=column + 2).border = Border(right=self.THICK)
                        else:
                            sheet.cell(row=student_index + 4, column=column + 2).border = Border(right=self.DOUBLE)

                subject_index += 1

            student_index += 1

        if len(wrong_marks) != 0:
            res_sheet = workbook.create_sheet('Results')

            wrong_marks.sort(key=lambda el: el['subject'])

            col_names = ('Ученик', 'Предмет', 'Период', 'Ср. б.', 'Рек.', 'Фактич.')
            for col in range(1, 7):
                res_sheet.cell(row=1, column=col, value=col_names[col - 1]).alignment = Alignment(horizontal='center')
                res_sheet.cell(row=1, column=col).font = Font(b=True)
                res_sheet.cell(row=1, column=col).border = Border(right=self.DOUBLE, bottom=self.DOUBLE)

            res_sheet.column_dimensions['A'].width = 35
            res_sheet.column_dimensions['B'].width = 25
            res_sheet.column_dimensions['C'].width = 20

            if int(form.split('-')[0]) in (10, 11):
                for i in range(len(wrong_marks)):
                    wrong_marks[i]['period'] = '{}-е полугодие'.format(wrong_marks[i]['period'] + 1)
            else:
                for i in range(len(wrong_marks)):
                    wrong_marks[i]['period'] = '{}-й триместр'.format(wrong_marks[i]['period'] + 1)

            keys = ('name', 'subject', 'period', 'average', 'recommended', 'actual')
            last = False
            for row in range(1, len(wrong_marks) + 1):
                if row == len(wrong_marks):
                    last = True
                for col in range(1, 7):
                    if not last:
                        res_sheet.cell(row=row + 1, column=col, value=wrong_marks[row - 1][keys[col - 1]])\
                            .border = Border(right=self.THIN)
                    else:
                        res_sheet.cell(row=row + 1, column=col, value=wrong_marks[row - 1][keys[col - 1]])\
                            .border = Border(right=self.THIN, bottom=self.THIN)

        workbook.save(filename)


def main():
    try:
        app = QApplication(sys.argv)
        gui = ExcelMarksInterface()
        sys.exit(app.exec())
    except Exception as e:
        print('Ошибка:', e)


if __name__ == '__main__':
    main()

